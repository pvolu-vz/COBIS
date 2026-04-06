#!/usr/bin/env python3
"""
COBIS to Veza OAA Integration Script

Reads a COBIS "Users By Group" Excel report (.xls/.xlsx) from a local or
SMB-mounted path and pushes user-to-group membership data into Veza's
Authorization Graph via the Open Authorization API (OAA).

Each sheet in the workbook represents one security group.  Users listed on
that sheet are members of the group.  The script de-duplicates users across
sheets and models:

  - Local Users   (one per unique User Name)
  - Local Groups  (one per sheet / group name)
  - Group membership (user ↔ group)

No permission mapping is performed — only identity-to-group relationships.
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
import tempfile
from datetime import datetime, timedelta, timezone

from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Optional SMB support — loaded lazily only when --smb-server is used
# ---------------------------------------------------------------------------
smbclient = None  # will be imported on demand

log = logging.getLogger(__name__)

# ── Excel column indices (0-based) matching the COBIS report layout ────────
COL_USERNAME = 0
COL_FIRST_NAME = 9
COL_DEPARTMENT = 12
COL_PLANT = 15
COL_DISABLED = 18
COL_DISABLED_DATE = 19
COL_INACTIVE = 22
COL_INACTIVATION_DATE = 24
COL_CREATED_DATE = 27
COL_CREATED_BY = 28
COL_LAST_MODIFIED = 29
COL_MODIFY_BY = 31

# Expected row containing the header labels (0-based); used as center of search window
HEADER_ROW = 12
# Group name row
GROUP_ROW = 13
# First data row
DATA_START_ROW = 15

# Excel epoch for serial-date conversion
EXCEL_EPOCH = datetime(1899, 12, 30, tzinfo=timezone.utc)


# ═══════════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════════

def _excel_date_to_iso(serial: float) -> str:
    """Convert an Excel serial date number to an ISO-8601 string."""
    if not serial or serial <= 0:
        return ""
    try:
        dt = EXCEL_EPOCH + timedelta(days=serial)
        return dt.strftime("%Y-%m-%dT%H:%M:%SZ")
    except (OverflowError, ValueError, OSError):
        return ""


def _str_val(cell_value) -> str:
    """Return a stripped string from a cell value, or empty string."""
    if cell_value is None:
        return ""
    return str(cell_value).strip()


def _bool_val(cell_value) -> bool:
    """Interpret a boolean / truthy cell value."""
    if isinstance(cell_value, bool):
        return cell_value
    if isinstance(cell_value, (int, float)):
        return bool(cell_value)
    if isinstance(cell_value, str):
        return cell_value.strip().lower() in ("true", "1", "yes", "x")
    return False


# ═══════════════════════════════════════════════════════════════════════════
# Excel parsing
# ═══════════════════════════════════════════════════════════════════════════

def _find_header_row(cell_getter, max_row: int, col: int = COL_USERNAME,
                     target: str = "user name") -> int:
    """Scan downward to find the row containing the header label.

    Args:
        cell_getter: callable(row, col) → cell value
        max_row: maximum row index to scan
        col: column index to check
        target: lowercase header text to match

    Returns:
        Row index of the header, or -1 if not found.
    """
    for r in range(max(0, HEADER_ROW - 2), min(max_row, HEADER_ROW + 5)):
        if _str_val(cell_getter(r, col)).lower() == target:
            return r
    return -1


def _read_xls(filepath: str) -> list[dict]:
    """Parse an old-format .xls workbook using xlrd and return records."""
    import xlrd

    wb = xlrd.open_workbook(filepath)
    records = []

    for name in wb.sheet_names():
        sh = wb.sheet_by_name(name)
        if sh.nrows < 15:
            log.debug("Skipping sheet '%s' — not enough rows (%d)", name, sh.nrows)
            continue

        # Dynamically locate the header row ("User Name" in col A)
        header_row = _find_header_row(sh.cell_value, sh.nrows)
        if header_row < 0:
            log.debug("Skipping sheet '%s' — could not find 'User Name' header", name)
            continue

        group_row = header_row + 1
        data_start = header_row + 3  # header → group → plant → first data row

        # Extract group name from the row after the header, col 9
        group_name = _str_val(sh.cell_value(group_row, COL_FIRST_NAME))
        if not group_name or group_name.lower().startswith("total"):
            log.warning("Sheet '%s' has no group name at row %d — skipping", name, group_row)
            continue

        log.info("Sheet '%s' → group '%s' (%d user rows)",
                 name, group_name, sh.nrows - data_start)

        for r in range(data_start, sh.nrows):
            username = _str_val(sh.cell_value(r, COL_USERNAME))
            if not username:
                continue

            records.append({
                "username": username,
                "first_name": _str_val(sh.cell_value(r, COL_FIRST_NAME)),
                "department": _str_val(sh.cell_value(r, COL_DEPARTMENT)),
                "plant": _str_val(sh.cell_value(r, COL_PLANT)),
                "is_disabled": _bool_val(sh.cell_value(r, COL_DISABLED)),
                "disabled_date": _excel_date_to_iso(
                    sh.cell_value(r, COL_DISABLED_DATE)
                    if sh.cell_type(r, COL_DISABLED_DATE) in (2, 3) else 0),
                "is_inactive": _bool_val(sh.cell_value(r, COL_INACTIVE)),
                "inactivation_date": _excel_date_to_iso(
                    sh.cell_value(r, COL_INACTIVATION_DATE)
                    if sh.cell_type(r, COL_INACTIVATION_DATE) in (2, 3) else 0),
                "created_date": _excel_date_to_iso(
                    sh.cell_value(r, COL_CREATED_DATE)
                    if sh.cell_type(r, COL_CREATED_DATE) in (2, 3) else 0),
                "created_by": _str_val(sh.cell_value(r, COL_CREATED_BY)),
                "last_modified": _excel_date_to_iso(
                    sh.cell_value(r, COL_LAST_MODIFIED)
                    if sh.cell_type(r, COL_LAST_MODIFIED) in (2, 3) else 0),
                "modified_by": _str_val(sh.cell_value(r, COL_MODIFY_BY)),
                "group": group_name,
            })

    return records


def _read_xlsx(filepath: str) -> list[dict]:
    """Parse an .xlsx workbook using openpyxl and return records."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    records = []

    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 15:
            log.debug("Skipping sheet '%s' — not enough rows", name)
            continue

        # Dynamically locate header row
        def _xlsx_cell(r, c):
            if r < len(rows) and c < len(rows[r]):
                return rows[r][c]
            return None

        header_row = _find_header_row(_xlsx_cell, len(rows))
        if header_row < 0:
            log.debug("Skipping sheet '%s' — header mismatch", name)
            continue

        group_row = header_row + 1
        data_start = header_row + 3

        group_name = _str_val(rows[group_row][COL_FIRST_NAME]) if len(rows[group_row]) > COL_FIRST_NAME else ""
        if not group_name:
            log.warning("Sheet '%s' has no group name — skipping", name)
            continue

        log.info("Sheet '%s' → group '%s' (%d user rows)",
                 name, group_name, len(rows) - data_start)

        for row in rows[data_start:]:
            if len(row) <= COL_USERNAME:
                continue
            username = _str_val(row[COL_USERNAME])
            if not username:
                continue

            def _cell(idx):
                return row[idx] if len(row) > idx else None

            def _date(idx):
                v = _cell(idx)
                if isinstance(v, (int, float)) and v > 0:
                    return _excel_date_to_iso(v)
                if isinstance(v, datetime):
                    return v.strftime("%Y-%m-%dT%H:%M:%SZ")
                return ""

            records.append({
                "username": username,
                "first_name": _str_val(_cell(COL_FIRST_NAME)),
                "department": _str_val(_cell(COL_DEPARTMENT)),
                "plant": _str_val(_cell(COL_PLANT)),
                "is_disabled": _bool_val(_cell(COL_DISABLED)),
                "disabled_date": _date(COL_DISABLED_DATE),
                "is_inactive": _bool_val(_cell(COL_INACTIVE)),
                "inactivation_date": _date(COL_INACTIVATION_DATE),
                "created_date": _date(COL_CREATED_DATE),
                "created_by": _str_val(_cell(COL_CREATED_BY)),
                "last_modified": _date(COL_LAST_MODIFIED),
                "modified_by": _str_val(_cell(COL_MODIFY_BY)),
                "group": group_name,
            })

    wb.close()
    return records


def read_excel(filepath: str) -> list[dict]:
    """Dispatch to the correct reader based on file extension."""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".xls":
        return _read_xls(filepath)
    elif ext in (".xlsx", ".xlsm"):
        return _read_xlsx(filepath)
    else:
        log.error("Unsupported file extension '%s' — expected .xls or .xlsx", ext)
        sys.exit(1)


def read_last_logon_xlsx(filepath: str) -> dict[str, str]:
    """Parse a T27991126A-style XLSX and return {user_code: last_logon_iso}.

    The file has a single sheet with a header row at row 4:
      C1=Groepsnaam  C2=Naam gebruiker  C3=Gebruikerscode  …  C8=Laatste login
    Data rows start at row 5.  A user may appear on multiple rows (one per
    group); the most recent last-logon value is kept.
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    logon_map: dict[str, datetime | None] = {}

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        # Locate header row (look for "Gebruikerscode" in the first 10 rows)
        header_idx = -1
        user_code_col = -1
        last_logon_col = -1
        for idx, row in enumerate(rows[:10]):
            for ci, cell in enumerate(row or []):
                val = _str_val(cell).lower()
                if "gebruikerscode" in val:
                    header_idx = idx
                    user_code_col = ci
                if "laatste login" in val or "recentste inlogmoment" in val:
                    last_logon_col = ci
            if header_idx >= 0:
                break

        if header_idx < 0 or user_code_col < 0 or last_logon_col < 0:
            log.warning("Sheet '%s' in last-logon file — could not locate headers; skipping", ws.title)
            continue

        log.info("Last-logon sheet '%s': header at row %d, user_code=C%d, last_logon=C%d",
                 ws.title, header_idx + 1, user_code_col + 1, last_logon_col + 1)

        for row in rows[header_idx + 1:]:
            if not row or len(row) <= max(user_code_col, last_logon_col):
                continue
            user_code = _str_val(row[user_code_col])
            if not user_code:
                continue

            raw = row[last_logon_col]
            if raw is None:
                continue

            # Parse into datetime for comparison
            if isinstance(raw, datetime):
                dt = raw
            elif isinstance(raw, (int, float)) and raw > 0:
                dt = EXCEL_EPOCH + timedelta(days=raw)
            else:
                continue

            existing = logon_map.get(user_code)
            if existing is None or dt > existing:
                logon_map[user_code] = dt

    wb.close()

    # Convert to ISO-8601 strings
    result: dict[str, str] = {}
    for code, dt in logon_map.items():
        if dt is not None:
            result[code] = dt.strftime("%Y-%m-%dT%H:%M:%SZ")

    log.info("Parsed last-logon data for %d users from %s", len(result), filepath)
    return result


# ═══════════════════════════════════════════════════════════════════════════
# SMB file retrieval
# ═══════════════════════════════════════════════════════════════════════════

def fetch_from_smb(server: str, share: str, path: str,
                   username: str, password: str, port: int = 445) -> str:
    """Download the Excel file from an SMB share to a local temp file.

    Returns the path to the local temporary copy.
    """
    global smbclient
    if smbclient is None:
        try:
            import smbclient as _smbclient
            smbclient = _smbclient
        except ImportError:
            log.error("smbprotocol is required for direct SMB access. "
                      "Install it with: pip install smbprotocol")
            sys.exit(1)

    smbclient.register_session(server, username=username, password=password, port=port)
    smb_path = f"\\\\{server}\\{share}\\{path.replace('/', os.sep)}"
    log.info("Downloading from SMB: %s", smb_path)

    ext = os.path.splitext(path)[1] or ".xls"
    tmp = tempfile.NamedTemporaryFile(suffix=ext, delete=False)
    try:
        with smbclient.open_file(smb_path, mode="rb") as remote:
            while True:
                chunk = remote.read(1024 * 1024)
                if not chunk:
                    break
                tmp.write(chunk)
        tmp.close()
        log.info("Downloaded %s to %s", smb_path, tmp.name)
        return tmp.name
    except Exception:
        tmp.close()
        os.unlink(tmp.name)
        raise


# ═══════════════════════════════════════════════════════════════════════════
# OAA payload assembly
# ═══════════════════════════════════════════════════════════════════════════

def build_oaa_payload(records: list[dict], datasource_name: str,
                      provider_name: str,
                      last_logon_map: dict[str, str] | None = None):
    """Build a Veza CustomApplication from the parsed records."""
    from oaaclient.templates import CustomApplication, OAAPropertyType

    app = CustomApplication(name=datasource_name, application_type=provider_name)

    # ── Define custom properties on Local Users ───────────────────────────
    app.property_definitions.define_local_user_property("department", OAAPropertyType.STRING)
    app.property_definitions.define_local_user_property("plant", OAAPropertyType.STRING)
    app.property_definitions.define_local_user_property("is_disabled", OAAPropertyType.BOOLEAN)
    app.property_definitions.define_local_user_property("disabled_date", OAAPropertyType.STRING)
    app.property_definitions.define_local_user_property("is_inactive", OAAPropertyType.BOOLEAN)
    app.property_definitions.define_local_user_property("inactivation_date", OAAPropertyType.STRING)
    app.property_definitions.define_local_user_property("created_date", OAAPropertyType.TIMESTAMP)
    app.property_definitions.define_local_user_property("created_by", OAAPropertyType.STRING)
    app.property_definitions.define_local_user_property("last_modified", OAAPropertyType.TIMESTAMP)
    app.property_definitions.define_local_user_property("modified_by", OAAPropertyType.STRING)
    app.property_definitions.define_local_user_property("last_login", OAAPropertyType.TIMESTAMP)

    groups_seen: set[str] = set()
    users_seen: dict[str, dict] = {}  # username → first record seen

    for rec in records:
        group = rec["group"]
        username = rec["username"]

        # Create group if new
        if group not in groups_seen:
            app.add_local_group(group)
            groups_seen.add(group)
            log.debug("Added group: %s", group)

        # Create user if new (first occurrence wins for attributes)
        if username not in users_seen:
            user = app.add_local_user(unique_id=username, name=rec["first_name"])
            user.is_active = not (rec["is_disabled"] or rec["is_inactive"])

            user.set_property("department", rec["department"])
            user.set_property("plant", rec["plant"])
            user.set_property("is_disabled", rec["is_disabled"])
            user.set_property("disabled_date", rec["disabled_date"])
            user.set_property("is_inactive", rec["is_inactive"])
            user.set_property("inactivation_date", rec["inactivation_date"])
            if rec["created_date"]:
                user.set_property("created_date", rec["created_date"])
                user.created_at = rec["created_date"]
            if rec["created_by"]:
                user.set_property("created_by", rec["created_by"])
            if rec["last_modified"]:
                user.set_property("last_modified", rec["last_modified"])
            if rec["modified_by"]:
                user.set_property("modified_by", rec["modified_by"])

            # Merge last-logon data from T27991126A report if available
            if last_logon_map:
                logon_ts = last_logon_map.get(username)
                if logon_ts:
                    user.last_login_at = logon_ts
                    user.set_property("last_login", logon_ts)

            users_seen[username] = rec

        # Assign user to group
        app.local_users[username].add_group(group)

    log.info("OAA payload built — %d users, %d groups, %d memberships",
             len(users_seen), len(groups_seen), len(records))

    return app


# ═══════════════════════════════════════════════════════════════════════════
# Push to Veza
# ═══════════════════════════════════════════════════════════════════════════

def push_to_veza(veza_url: str, veza_api_key: str, provider_name: str,
                 datasource_name: str, app, dry_run: bool = False,
                 save_json: bool = False):
    """Push the OAA application payload to Veza."""
    from oaaclient.client import OAAClient, OAAClientError

    if dry_run:
        log.info("[DRY RUN] Payload built successfully — skipping push to Veza")
        if save_json:
            import json
            payload = app.get_payload()
            out_path = f"{datasource_name.replace(' ', '_')}_payload.json"
            with open(out_path, "w") as f:
                json.dump(payload, f, indent=2, default=str)
            log.info("[DRY RUN] Payload saved to %s", out_path)
        return

    veza_con = OAAClient(url=veza_url, token=veza_api_key)
    log.info("Connected to Veza at %s", veza_url)

    provider = veza_con.get_provider(provider_name)
    if provider:
        log.info("Found existing provider '%s'", provider_name)
    else:
        log.info("Creating provider '%s'", provider_name)
        provider = veza_con.create_provider(provider_name, "application")
    log.info("Provider: %s (%s)", provider["name"], provider["id"])

    try:
        response = veza_con.push_application(
            provider_name=provider_name,
            data_source_name=datasource_name,
            application_object=app,
            save_json=save_json,
        )
        if response.get("warnings"):
            for w in response["warnings"]:
                logging.warning("Veza warning: %s", w)
        log.info("Successfully pushed to Veza")
    except OAAClientError as e:
        log.error("Veza push failed: %s — %s (HTTP %s)", e.error, e.message, e.status_code)
        if hasattr(e, "details"):
            for d in e.details:
                log.error("  Detail: %s", d)
        sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════════
# Configuration
# ═══════════════════════════════════════════════════════════════════════════

def load_config(args):
    """Merge CLI args, env vars, and .env file. CLI args take precedence."""
    if args.env_file and os.path.exists(args.env_file):
        load_dotenv(args.env_file)
        log.info("Loaded env file: %s", args.env_file)

    return {
        "veza_url": args.veza_url or os.getenv("VEZA_URL"),
        "veza_api_key": args.veza_api_key or os.getenv("VEZA_API_KEY"),
        "provider_name": args.provider_name or os.getenv("PROVIDER_NAME", "COBIS"),
        "datasource_name": args.datasource_name or os.getenv("DATASOURCE_NAME", "COBIS"),
        "xlsx_path": args.xlsx_path or os.getenv("COBIS_XLSX_PATH"),
        "smb_server": args.smb_server or os.getenv("COBIS_SMB_SERVER"),
        "smb_share": args.smb_share or os.getenv("COBIS_SMB_SHARE"),
        "smb_path": args.smb_path or os.getenv("COBIS_SMB_PATH"),
        "smb_username": args.smb_username or os.getenv("COBIS_SMB_USERNAME"),
        "smb_password": args.smb_password or os.getenv("COBIS_SMB_PASSWORD"),
        "smb_port": args.smb_port or int(os.getenv("COBIS_SMB_PORT", "445")),
        "last_logon_path": args.last_logon_path or os.getenv("COBIS_LAST_LOGON_PATH"),
        "dry_run": args.dry_run,
        "save_json": args.save_json,
    }


# ═══════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="COBIS → Veza OAA integration: push user-to-group data from a "
                    "COBIS 'Users By Group' Excel report into Veza.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    # ── General ────────────────────────────────────────────────────────────
    parser.add_argument("--env-file", default=".env",
                        help="Path to .env file (default: .env)")
    parser.add_argument("--log-level", default="INFO",
                        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
                        help="Logging level (default: INFO)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Build payload but skip push to Veza")
    parser.add_argument("--save-json", action="store_true",
                        help="Save OAA JSON payload to a local file")

    # ── Veza ───────────────────────────────────────────────────────────────
    veza = parser.add_argument_group("Veza connection")
    veza.add_argument("--veza-url", default=None,
                      help="Veza tenant URL (env: VEZA_URL)")
    veza.add_argument("--veza-api-key", default=None,
                      help="Veza API key (env: VEZA_API_KEY)")
    veza.add_argument("--provider-name", default=None,
                      help="Veza provider name (default: COBIS)")
    veza.add_argument("--datasource-name", default=None,
                      help="Veza datasource name (default: COBIS)")

    # ── Excel source (local / mounted) ─────────────────────────────────────
    src = parser.add_argument_group("Excel source (local or mounted path)")
    src.add_argument("--xlsx-path", default=None,
                     help="Path to the COBIS .xls/.xlsx file (env: COBIS_XLSX_PATH)")
    src.add_argument("--last-logon-path", default=None,
                     help="Path to the T27991126A last-logon .xlsx file "
                          "(env: COBIS_LAST_LOGON_PATH)")

    # ── SMB direct access ──────────────────────────────────────────────────
    smb = parser.add_argument_group("SMB direct access (alternative to --xlsx-path)")
    smb.add_argument("--smb-server", default=None,
                     help="SMB/CIFS server hostname or IP (env: COBIS_SMB_SERVER)")
    smb.add_argument("--smb-share", default=None,
                     help="SMB share name (env: COBIS_SMB_SHARE)")
    smb.add_argument("--smb-path", default=None,
                     help="Path to file within the share (env: COBIS_SMB_PATH)")
    smb.add_argument("--smb-username", default=None,
                     help="SMB username (env: COBIS_SMB_USERNAME)")
    smb.add_argument("--smb-password", default=None,
                     help="SMB password (env: COBIS_SMB_PASSWORD)")
    smb.add_argument("--smb-port", type=int, default=None,
                     help="SMB port (default: 445, env: COBIS_SMB_PORT)")

    return parser


# ═══════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════

def main():
    parser = build_parser()
    args = parser.parse_args()

    # ── Logging ────────────────────────────────────────────────────────────
    logging.basicConfig(
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        level=getattr(logging, args.log_level),
    )

    # Startup banner (only bare print allowed)
    print("=" * 60)
    print("  COBIS → Veza OAA Integration")
    print("=" * 60)

    cfg = load_config(args)

    # ── Validate required Veza config ──────────────────────────────────────
    if not cfg["dry_run"]:
        missing = []
        if not cfg["veza_url"]:
            missing.append("--veza-url / VEZA_URL")
        if not cfg["veza_api_key"]:
            missing.append("--veza-api-key / VEZA_API_KEY")
        if missing:
            log.error("Missing required Veza parameters: %s", ", ".join(missing))
            sys.exit(1)

    # ── Resolve Excel file ─────────────────────────────────────────────────
    local_path = cfg["xlsx_path"]
    smb_tmp = None

    if not local_path and cfg["smb_server"]:
        # Direct SMB download
        smb_missing = []
        for key in ("smb_server", "smb_share", "smb_path", "smb_username", "smb_password"):
            if not cfg[key]:
                smb_missing.append(f"--{key.replace('_', '-')} / COBIS_{key.upper()}")
        if smb_missing:
            log.error("Missing SMB parameters: %s", ", ".join(smb_missing))
            sys.exit(1)
        local_path = fetch_from_smb(
            server=cfg["smb_server"],
            share=cfg["smb_share"],
            path=cfg["smb_path"],
            username=cfg["smb_username"],
            password=cfg["smb_password"],
            port=cfg["smb_port"],
        )
        smb_tmp = local_path  # remember to clean up

    if not local_path:
        log.error("No Excel file specified. Use --xlsx-path or --smb-server/share/path.")
        sys.exit(1)

    if not os.path.isfile(local_path):
        log.error("File not found: %s", local_path)
        sys.exit(1)

    # ── Parse Excel ────────────────────────────────────────────────────────
    log.info("Reading Excel file: %s", local_path)
    try:
        records = read_excel(local_path)
    finally:
        if smb_tmp and os.path.exists(smb_tmp):
            os.unlink(smb_tmp)
            log.debug("Cleaned up temp file %s", smb_tmp)

    if not records:
        log.error("No user records found in the Excel file")
        sys.exit(1)

    log.info("Parsed %d user-group records from Excel", len(records))

    # ── Parse last-logon report (optional) ─────────────────────────────────
    last_logon_map = None
    if cfg["last_logon_path"]:
        ll_path = cfg["last_logon_path"]
        if not os.path.isfile(ll_path):
            log.error("Last-logon file not found: %s", ll_path)
            sys.exit(1)
        log.info("Reading last-logon file: %s", ll_path)
        last_logon_map = read_last_logon_xlsx(ll_path)

    # ── Build OAA payload ──────────────────────────────────────────────────
    app = build_oaa_payload(records, cfg["datasource_name"], cfg["provider_name"],
                            last_logon_map=last_logon_map)

    # ── Push to Veza ───────────────────────────────────────────────────────
    push_to_veza(
        veza_url=cfg["veza_url"],
        veza_api_key=cfg["veza_api_key"],
        provider_name=cfg["provider_name"],
        datasource_name=cfg["datasource_name"],
        app=app,
        dry_run=cfg["dry_run"],
        save_json=cfg["save_json"],
    )

    log.info("Run finished")


if __name__ == "__main__":
    main()
